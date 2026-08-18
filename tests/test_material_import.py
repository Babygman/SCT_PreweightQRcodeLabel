import json
import zipfile
from io import BytesIO
from pathlib import Path

import pytest
from openpyxl import Workbook
from sqlalchemy.dialects import mssql
from werkzeug.security import generate_password_hash

from app.extensions import db
from app.models import (
    AuditLog,
    Material,
    MaterialImportBatch,
    MaterialImportRow,
    MaterialTag,
    MaterialTagBatch,
    MaterialTagDraft,
    Role,
    Station,
    User,
)
from app.services.material_import import (
    MaterialImportError,
    _batch_for_apply_statement,
    apply_material_import,
    parse_material_workbook,
)
from config import Config, DevelopmentConfig, TestingConfig

REAL_WORKBOOK = Path("/Users/rachin/Downloads/Material code.xlsx")


def test_material_import_feature_gate_is_disabled_in_all_default_configurations():
    assert Config.MATERIAL_TAG_ISSUANCE_ENABLED is False
    assert DevelopmentConfig.MATERIAL_TAG_ISSUANCE_ENABLED is False
    assert TestingConfig.MATERIAL_TAG_ISSUANCE_ENABLED is False


def _identity(app, role_code="ADMIN"):
    with app.app_context():
        role = Role(code=role_code, name=role_code.title())
        user = User(
            username=f"{role_code.lower()}-importer",
            password_hash=generate_password_hash("Test-only-password!"),
            display_name=f"{role_code.title()} Importer",
            roles=[role],
        )
        station = Station(code=f"{role_code[:3]}-ST", name="Import Station")
        db.session.add_all([role, user, station])
        db.session.commit()
        return user.id, station.id


def _authenticate(client, user_id, station_id):
    with client.session_transaction() as user_session:
        user_session["_user_id"] = str(user_id)
        user_session["_fresh"] = True
        user_session["station_id"] = station_id


def _workbook_bytes(rows, headers=("ITEM CODE", "CATEGORY_NO", "NAME")):
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Sheet1"
    worksheet.append(headers)
    for row in rows:
        worksheet.append(row)
    output = BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


def _upload(
    client,
    workbook_bytes,
    idempotency_key="11111111-1111-4111-8111-111111111111",
    filename="materials.xlsx",
):
    return client.post(
        "/master-data/materials/import",
        data={
            "idempotency_key": idempotency_key,
            "workbook": (BytesIO(workbook_bytes), filename),
        },
        content_type="multipart/form-data",
        follow_redirects=False,
    )


def _rewrite_xlsx(workbook_bytes, *, replacements=None, additions=None):
    replacements = replacements or {}
    additions = additions or {}
    output = BytesIO()
    with (
        zipfile.ZipFile(BytesIO(workbook_bytes)) as source,
        zipfile.ZipFile(output, "w", zipfile.ZIP_DEFLATED) as destination,
    ):
        for member in source.infolist():
            content = replacements.get(member.filename, source.read(member))
            destination.writestr(member, content)
        for name, content in additions.items():
            destination.writestr(name, content)
    return output.getvalue()


def test_feature_is_disabled_by_default_without_querying_new_tables(app, client):
    user_id, station_id = _identity(app)
    _authenticate(client, user_id, station_id)
    with app.app_context():
        MaterialImportRow.__table__.drop(db.engine)
        MaterialImportBatch.__table__.drop(db.engine)
    assert client.get("/master-data/materials/import").status_code == 404
    assert client.get("/master-data/materials/import/1/preview").status_code == 404
    assert client.post("/master-data/materials/import/1/apply").status_code == 404
    assert client.get("/master-data/materials/import/1/result").status_code == 404
    master_data = client.get("/master-data/")
    assert master_data.status_code == 200
    assert b"Import Material Master" not in master_data.data


@pytest.mark.parametrize("role_code", ["OPERATOR", "PRODUCTION", "SUPERVISOR"])
def test_material_import_routes_are_admin_only(app, client, role_code):
    app.config["MATERIAL_TAG_ISSUANCE_ENABLED"] = True
    user_id, station_id = _identity(app, role_code)
    _authenticate(client, user_id, station_id)
    assert client.get("/master-data/materials/import").status_code == 403
    assert client.get("/master-data/materials/import/1/preview").status_code == 403
    assert client.post("/master-data/materials/import/1/apply").status_code == 403
    assert client.get("/master-data/materials/import/1/result").status_code == 403


def test_material_import_requires_authentication_and_station(app, client):
    app.config["MATERIAL_TAG_ISSUANCE_ENABLED"] = True
    for method, path in (
        (client.get, "/master-data/materials/import"),
        (client.get, "/master-data/materials/import/1/preview"),
        (client.post, "/master-data/materials/import/1/apply"),
        (client.get, "/master-data/materials/import/1/result"),
    ):
        assert method(path).status_code == 302
    _identity(app)
    client.post(
        "/auth/login",
        data={"username": "admin-importer", "password": "Test-only-password!"},
    )
    response = client.get("/master-data/materials/import")
    assert response.status_code == 302
    assert "/auth/station" in response.headers["Location"]


def test_apply_requires_csrf_when_enabled(app, client):
    app.config.update(MATERIAL_TAG_ISSUANCE_ENABLED=True, WTF_CSRF_ENABLED=True)
    user_id, station_id = _identity(app)
    _authenticate(client, user_id, station_id)
    assert client.post("/master-data/materials/import/1/apply").status_code == 400
    assert (
        client.post(
            "/master-data/materials/import",
            data={
                "idempotency_key": "11111111-1111-4111-8111-111111111111",
                "workbook": (BytesIO(_workbook_bytes([("MAT-1", "MAT", "Name")])), "a.xlsx"),
            },
            content_type="multipart/form-data",
        ).status_code
        == 400
    )


def test_structural_failure_is_persistent_and_audited(app, client):
    app.config["MATERIAL_TAG_ISSUANCE_ENABLED"] = True
    user_id, station_id = _identity(app)
    _authenticate(client, user_id, station_id)
    response = _upload(client, b"not an xlsx")
    assert response.status_code == 302
    with app.app_context():
        batch = MaterialImportBatch.query.one()
        assert batch.status == "FAILED"
        assert "valid .xlsx" in batch.error_summary
        audit = AuditLog.query.filter_by(event_type="MATERIAL_IMPORT_FAILED").one()
        detail = json.loads(audit.detail)
        assert detail["batch_id"] == batch.id
        assert detail["filename"] == "materials.xlsx"
        assert detail["file_sha256"] == batch.file_sha256
        assert detail["status"] == "FAILED"
        assert "password" not in audit.detail.lower()


@pytest.mark.parametrize("filename", ["materials.xls", "materials.xlsm", "materials.txt"])
def test_upload_rejects_non_xlsx_filename(app, client, filename):
    app.config["MATERIAL_TAG_ISSUANCE_ENABLED"] = True
    user_id, station_id = _identity(app)
    _authenticate(client, user_id, station_id)
    response = _upload(
        client,
        _workbook_bytes([("MAT-1", "MAT", "Material")]),
        filename=filename,
    )
    assert response.status_code == 200
    assert b"Only .xlsx workbooks are accepted" in response.data
    with app.app_context():
        assert MaterialImportBatch.query.count() == 0


def test_xlsx_container_security_rejections():
    base = _workbook_bytes([("MAT-1", "MAT", "Material")])
    external = _rewrite_xlsx(
        base,
        additions={
            "xl/externalLinks/_rels/externalLink1.xml.rels": (
                b'<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/'
                b'relationships"><Relationship Id="rId1" Target="https://example.invalid" '
                b'TargetMode="External" Type="external"/></Relationships>'
            )
        },
    )
    macro = _rewrite_xlsx(base, additions={"xl/vbaProject.bin": b"not executed"})
    traversal = _rewrite_xlsx(base, additions={"../unsafe.xml": b"<unsafe/>"})
    malformed = _rewrite_xlsx(base, replacements={"xl/workbook.xml": b"<workbook>"})
    with zipfile.ZipFile(BytesIO(base)) as archive:
        sheet_xml = archive.read("xl/worksheets/sheet1.xml")
    merged = _rewrite_xlsx(
        base,
        replacements={
            "xl/worksheets/sheet1.xml": sheet_xml.replace(
                b"</worksheet>",
                b'<mergeCells count="1"><mergeCell ref="A2:B2"/></mergeCells></worksheet>',
            )
        },
    )
    cases = (
        (external, "external links"),
        (macro, "Macro-enabled"),
        (traversal, "unsafe internal path"),
        (malformed, "malformed or unsafe XML"),
        (merged, "Merged cells"),
        (b"PK renamed plain text", "valid .xlsx"),
        (b"\xd0\xcf\x11\xe0encrypted office", "valid .xlsx"),
    )
    for workbook_bytes, message in cases:
        with pytest.raises(MaterialImportError, match=message):
            parse_material_workbook(
                workbook_bytes,
                maximum_bytes=5 * 1024 * 1024,
                maximum_rows=5_000,
                maximum_uncompressed_bytes=50 * 1024 * 1024,
            )


def test_xlsx_container_enforces_size_entry_and_expansion_limits():
    base = _workbook_bytes([("MAT-1", "MAT", "Material")])
    with pytest.raises(MaterialImportError, match="upload size"):
        parse_material_workbook(
            base + (b"x" * (5 * 1024 * 1024)),
            maximum_bytes=5 * 1024 * 1024,
            maximum_rows=5_000,
            maximum_uncompressed_bytes=50 * 1024 * 1024,
        )
    too_many_entries = _rewrite_xlsx(
        base,
        additions={f"xl/unused/{index}.bin": b"x" for index in range(1_001)},
    )
    with pytest.raises(MaterialImportError, match="too many internal files"):
        parse_material_workbook(
            too_many_entries,
            maximum_bytes=5 * 1024 * 1024,
            maximum_rows=5_000,
            maximum_uncompressed_bytes=50 * 1024 * 1024,
        )
    expanded = _rewrite_xlsx(base, additions={"xl/unused/padding.bin": b"x" * 1_000})
    with pytest.raises(MaterialImportError, match="expands beyond"):
        parse_material_workbook(
            expanded,
            maximum_bytes=5 * 1024 * 1024,
            maximum_rows=5_000,
            maximum_uncompressed_bytes=500,
        )


def test_filename_is_sanitized_and_workbook_binary_is_not_persisted(app, client):
    app.config["MATERIAL_TAG_ISSUANCE_ENABLED"] = True
    user_id, station_id = _identity(app)
    _authenticate(client, user_id, station_id)
    response = _upload(
        client,
        _workbook_bytes([("MAT-1", "MAT", "Material")]),
        filename="../../private/materials.xlsx",
    )
    page = client.get(response.headers["Location"])
    assert b"private/" not in page.data
    assert b"materials.xlsx" in page.data
    with app.app_context():
        batch = MaterialImportBatch.query.one()
        assert batch.original_filename == "materials.xlsx"
        assert "binary" not in MaterialImportBatch.__table__.columns
        assert "content" not in MaterialImportBatch.__table__.columns


def test_row_validation_normalization_and_duplicate_classification():
    workbook_bytes = _workbook_bytes(
        [
            (" mat-1 ", " mat ", " Name One "),
            ("MAT-1", "MAT", "Duplicate Code"),
            (None, "MAT", "Missing Code"),
            ("MAT-2", "OTHER", "Wrong Category"),
            ("MAT-3", "MAT", "=1+1"),
        ]
    )
    rows = parse_material_workbook(
        workbook_bytes,
        maximum_bytes=5 * 1024 * 1024,
        maximum_rows=10_000,
        maximum_uncompressed_bytes=50 * 1024 * 1024,
    )
    assert [row.reason_code for row in rows] == [
        "DUPLICATE_ITEM_CODE",
        "DUPLICATE_ITEM_CODE",
        "REQUIRED_VALUE_MISSING",
        "CATEGORY_NOT_ALLOWED",
        "FORMULA_NOT_ALLOWED",
    ]
    assert rows[0].item_code == "MAT-1"
    assert rows[0].name == "Name One"


@pytest.mark.parametrize(
    "row",
    [("=1+1", "MAT", "Name"), ("MAT-1", "=1+1", "Name"), ("MAT-1", "MAT", "=1+1")],
)
def test_formula_in_any_required_business_cell_is_rejected(row):
    parsed = parse_material_workbook(
        _workbook_bytes([row]),
        maximum_bytes=5 * 1024 * 1024,
        maximum_rows=5_000,
        maximum_uncompressed_bytes=50 * 1024 * 1024,
    )
    assert parsed[0].reason_code == "FORMULA_NOT_ALLOWED"


def test_control_characters_are_rejected_at_row_level():
    rows = parse_material_workbook(
        _workbook_bytes([("MAT-1", "MAT", "Unsafe\nName"), ("MAT-2", "MAT", "Safe")]),
        maximum_bytes=5 * 1024 * 1024,
        maximum_rows=5_000,
        maximum_uncompressed_bytes=50 * 1024 * 1024,
    )
    assert rows[0].row_number == 2
    assert rows[0].item_code == "MAT-1"
    assert rows[0].category_no == "MAT"
    assert rows[0].reason_code == "CONTROL_CHARACTER_NOT_ALLOWED"
    assert rows[1].row_number == 3
    assert rows[1].reason_code is None


def test_blank_trailing_rows_are_ignored_but_interior_blank_row_is_rejected():
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Sheet1"
    worksheet.append(("ITEM CODE", "CATEGORY_NO", "NAME"))
    worksheet.append(("MAT-1", "MAT", "One"))
    worksheet.append((None, None, None))
    worksheet.append(("MAT-2", "MAT", "Two"))
    worksheet.append((None, None, None))
    output = BytesIO()
    workbook.save(output)
    rows = parse_material_workbook(
        output.getvalue(),
        maximum_bytes=5 * 1024 * 1024,
        maximum_rows=5_000,
        maximum_uncompressed_bytes=50 * 1024 * 1024,
    )
    assert [row.row_number for row in rows] == [2, 3, 4]
    assert rows[1].reason_code == "REQUIRED_VALUE_MISSING"


def test_more_than_five_thousand_rows_and_empty_workbook_are_rejected():
    rows = [(f"MAT-{index:04d}", "MAT", f"Material {index}") for index in range(5_001)]
    with pytest.raises(MaterialImportError, match="more rows than permitted"):
        parse_material_workbook(
            _workbook_bytes(rows),
            maximum_bytes=5 * 1024 * 1024,
            maximum_rows=5_000,
            maximum_uncompressed_bytes=50 * 1024 * 1024,
        )
    with pytest.raises(MaterialImportError, match="no Material rows"):
        parse_material_workbook(
            _workbook_bytes([]),
            maximum_bytes=5 * 1024 * 1024,
            maximum_rows=5_000,
            maximum_uncompressed_bytes=50 * 1024 * 1024,
        )


def test_workbook_rejects_wrong_headers_and_extra_sheet():
    bad_headers = _workbook_bytes([("MAT-1", "MAT", "Name")], headers=("CODE", "CAT", "NAME"))
    with pytest.raises(MaterialImportError, match="headers must be exactly"):
        parse_material_workbook(
            bad_headers,
            maximum_bytes=5 * 1024 * 1024,
            maximum_rows=10_000,
            maximum_uncompressed_bytes=50 * 1024 * 1024,
        )

    trimmed_headers = parse_material_workbook(
        _workbook_bytes(
            [("MAT-1", "MAT", "Name")],
            headers=(" ITEM CODE ", " CATEGORY_NO ", " NAME "),
        ),
        maximum_bytes=5 * 1024 * 1024,
        maximum_rows=5_000,
        maximum_uncompressed_bytes=50 * 1024 * 1024,
    )
    assert trimmed_headers[0].item_code == "MAT-1"

    for headers in (
        ("ITEM CODE", "CATEGORY_NO"),
        ("ITEM CODE", "CATEGORY_NO", "NAME", "EXTRA"),
        ("ITEM CODE", "ITEM CODE", "NAME"),
    ):
        with pytest.raises(MaterialImportError, match="headers must be exactly"):
            parse_material_workbook(
                _workbook_bytes([("MAT-1", "MAT", "Name")], headers=headers),
                maximum_bytes=5 * 1024 * 1024,
                maximum_rows=5_000,
                maximum_uncompressed_bytes=50 * 1024 * 1024,
            )

    workbook = Workbook()
    workbook.active.title = "Sheet1"
    workbook.active.append(("ITEM CODE", "CATEGORY_NO", "NAME"))
    workbook.active.append(("MAT-1", "MAT", "Name"))
    workbook.create_sheet("Other")
    output = BytesIO()
    workbook.save(output)
    with pytest.raises(MaterialImportError, match="only the Sheet1"):
        parse_material_workbook(
            output.getvalue(),
            maximum_bytes=5 * 1024 * 1024,
            maximum_rows=10_000,
            maximum_uncompressed_bytes=50 * 1024 * 1024,
        )


@pytest.mark.skipif(
    not REAL_WORKBOOK.exists(), reason="Approved Material Master workbook unavailable"
)
def test_real_workbook_preview_apply_and_idempotency(app, client):
    app.config["MATERIAL_TAG_ISSUANCE_ENABLED"] = True
    user_id, station_id = _identity(app)
    _authenticate(client, user_id, station_id)
    with app.app_context():
        db.session.add_all(
            [
                Material(
                    code="R01034L2",
                    name="Silquest A-1110 (17kg/Pail)",
                    unit="g",
                    classification="SPECIAL",
                    source_category_no="MAT",
                    is_active=False,
                ),
                Material(
                    code="R01038L1",
                    name="Stale Name",
                    unit="lb",
                    classification="PRESERVE",
                    source_category_no="OLD",
                    is_active=False,
                ),
            ]
        )
        db.session.commit()

    response = _upload(client, REAL_WORKBOOK.read_bytes())
    assert response.status_code == 302
    assert response.headers["Location"].endswith("/master-data/materials/import/1/preview")
    preview = client.get(response.headers["Location"])
    assert preview.status_code == 200
    assert b"Preview only" in preview.data
    assert b"Next" in preview.data
    assert b"Thailand Time" in preview.data

    with app.app_context():
        assert Material.query.count() == 2
        batch = MaterialImportBatch.query.one()
        assert (batch.total_rows, batch.inserted_count, batch.updated_count) == (329, 327, 1)
        assert (batch.unchanged_count, batch.rejected_count) == (1, 0)
        pg740 = MaterialImportRow.query.filter_by(name_normalized="PG740").one()
        assert pg740.item_code_normalized == "R07047S1"
        assert pg740.category_no_normalized == "MAT"
        assert pg740.name_normalized == "PG740"
        assert len({row.item_code_normalized for row in batch.rows}) == 329
        assert len({row.name_normalized for row in batch.rows}) == 323
        assert {row.category_no_normalized for row in batch.rows} == {"MAT"}
        assert (
            batch.inserted_count
            + batch.updated_count
            + batch.unchanged_count
            + batch.rejected_count
            == 329
        )
        assert AuditLog.query.filter_by(event_type="MATERIAL_IMPORT_PREVIEWED").count() == 1

    applied = client.post("/master-data/materials/import/1/apply", follow_redirects=True)
    assert applied.status_code == 200
    assert b"Material Master import applied successfully" in applied.data
    with app.app_context():
        assert Material.query.count() == 329
        updated = Material.query.filter_by(code="R01038L1").one()
        assert updated.name != "Stale Name"
        assert (updated.unit, updated.classification, updated.is_active) == (
            "lb",
            "PRESERVE",
            False,
        )
        unchanged = Material.query.filter_by(code="R01034L2").one()
        assert (unchanged.unit, unchanged.classification, unchanged.is_active) == (
            "g",
            "SPECIAL",
            False,
        )
        assert MaterialImportBatch.query.one().status == "APPLIED"
        apply_audit = AuditLog.query.filter_by(event_type="MATERIAL_IMPORT_APPLIED").one()
        assert json.loads(apply_audit.detail)["filename"] == "materials.xlsx"
        assert MaterialTagDraft.query.count() == 0
        assert MaterialTagBatch.query.count() == 0
        assert MaterialTag.query.count() == 0

        applied_at = MaterialImportBatch.query.one().applied_at_utc

    repeated = client.post("/master-data/materials/import/1/apply")
    assert repeated.status_code == 302
    assert repeated.headers["Location"].endswith("/master-data/materials/import/1/result")
    with app.app_context():
        assert Material.query.count() == 329
        assert MaterialImportBatch.query.one().applied_at_utc == applied_at
        assert AuditLog.query.filter_by(event_type="MATERIAL_IMPORT_APPLIED").count() == 1

    second = _upload(
        client,
        REAL_WORKBOOK.read_bytes(),
        idempotency_key="22222222-2222-4222-8222-222222222222",
    )
    assert second.headers["Location"].endswith("/master-data/materials/import/2/preview")
    with app.app_context():
        second_batch = db.session.get(MaterialImportBatch, 2)
        assert (
            second_batch.inserted_count,
            second_batch.updated_count,
            second_batch.unchanged_count,
            second_batch.rejected_count,
        ) == (0, 0, 329, 0)
    client.post("/master-data/materials/import/2/apply")
    with app.app_context():
        assert Material.query.count() == 329
        assert AuditLog.query.filter_by(event_type="MATERIAL_IMPORT_APPLIED").count() == 2


def test_upsert_does_not_delete_or_deactivate_materials_absent_from_workbook(app, client):
    app.config["MATERIAL_TAG_ISSUANCE_ENABLED"] = True
    user_id, station_id = _identity(app)
    _authenticate(client, user_id, station_id)
    with app.app_context():
        db.session.add(
            Material(
                code="EXISTING-ABSENT",
                name="Not In Workbook",
                unit="g",
                classification="SPECIAL",
                source_category_no="LEGACY",
                is_active=True,
            )
        )
        db.session.commit()
    _upload(client, _workbook_bytes([("NEW-MAT", "MAT", "New Material")]))
    client.post("/master-data/materials/import/1/apply")
    with app.app_context():
        existing = Material.query.filter_by(code="EXISTING-ABSENT").one()
        assert existing.is_active is True
        assert (existing.name, existing.unit, existing.classification) == (
            "Not In Workbook",
            "g",
            "SPECIAL",
        )


def test_same_upload_idempotency_key_reuses_persistent_preview(app, client):
    app.config["MATERIAL_TAG_ISSUANCE_ENABLED"] = True
    user_id, station_id = _identity(app)
    _authenticate(client, user_id, station_id)
    workbook = _workbook_bytes([("MAT-1", "MAT", "Material One")])
    first = _upload(client, workbook)
    second = _upload(client, workbook)
    assert first.headers["Location"] == second.headers["Location"]
    with app.app_context():
        assert MaterialImportBatch.query.count() == 1
        assert MaterialImportRow.query.count() == 1
        assert AuditLog.query.filter_by(event_type="MATERIAL_IMPORT_PREVIEWED").count() == 1


def test_rejected_row_blocks_all_apply_including_forged_post(app, client):
    app.config["MATERIAL_TAG_ISSUANCE_ENABLED"] = True
    user_id, station_id = _identity(app)
    _authenticate(client, user_id, station_id)
    workbook = _workbook_bytes([("MAT-VALID", "MAT", "Valid"), ("MAT-BAD", "OTHER", "Rejected")])
    preview_response = _upload(client, workbook)
    preview = client.get(preview_response.headers["Location"])
    assert b"no Material record has changed yet" in preview.data
    assert b"Apply is blocked" in preview.data
    assert b"disabled" in preview.data

    response = client.post("/master-data/materials/import/1/apply", follow_redirects=True)
    assert response.status_code == 200
    assert b"contains rejected rows" in response.data
    with app.app_context():
        assert Material.query.count() == 0
        assert MaterialImportBatch.query.one().status == "PREVIEWED"
        assert AuditLog.query.filter_by(event_type="MATERIAL_IMPORT_APPLIED").count() == 0


def test_apply_revalidates_persisted_preview_integrity(app, client):
    app.config["MATERIAL_TAG_ISSUANCE_ENABLED"] = True
    user_id, station_id = _identity(app)
    _authenticate(client, user_id, station_id)
    _upload(client, _workbook_bytes([("MAT-1", "MAT", "Material")]))
    with app.app_context():
        row = MaterialImportRow.query.one()
        row.category_no_normalized = "OTHER"
        db.session.commit()
    response = client.post("/master-data/materials/import/1/apply", follow_redirects=True)
    assert b"persisted preview contains an invalid Material row" in response.data
    with app.app_context():
        assert Material.query.count() == 0
        assert MaterialImportBatch.query.one().status == "PREVIEWED"


def test_sql_server_apply_query_uses_update_and_serializable_range_locks():
    sql = str(_batch_for_apply_statement(1).compile(dialect=mssql.dialect()))
    assert "WITH (UPDLOCK, HOLDLOCK)" in sql


def test_apply_rolls_back_materials_batch_and_audit_together(app, monkeypatch):
    import app.services.material_import as service

    with app.app_context():
        user_id, station_id = _identity(app)
        batch = MaterialImportBatch(
            original_filename="materials.xlsx",
            file_sha256="0" * 64,
            status="PREVIEWED",
            total_rows=1,
            inserted_count=1,
            updated_count=0,
            unchanged_count=0,
            rejected_count=0,
            uploaded_by_user_id=user_id,
            uploaded_at_utc=service.utcnow(),
            idempotency_key="33333333-3333-4333-8333-333333333333",
        )
        batch.rows.append(
            MaterialImportRow(
                row_number=2,
                item_code_normalized="MAT-ROLLBACK",
                category_no_normalized="MAT",
                name_normalized="Rollback Material",
                result="INSERT",
            )
        )
        db.session.add(batch)
        db.session.commit()
        batch_id = batch.id

        original_add_audit = service._add_audit

        def fail_apply_audit(event_type, *args, **kwargs):
            if event_type == "MATERIAL_IMPORT_APPLIED":
                raise RuntimeError("forced isolated rollback")
            return original_add_audit(event_type, *args, **kwargs)

        monkeypatch.setattr(service, "_add_audit", fail_apply_audit)
        with pytest.raises(RuntimeError, match="forced isolated rollback"):
            apply_material_import(batch_id=batch_id, user_id=user_id, station_id=station_id)
        db.session.rollback()
        assert Material.query.count() == 0
        assert db.session.get(MaterialImportBatch, batch_id).status == "PREVIEWED"
        assert AuditLog.query.filter_by(event_type="MATERIAL_IMPORT_APPLIED").count() == 0


@pytest.mark.skipif(
    not REAL_WORKBOOK.exists(), reason="Approved Material Master workbook unavailable"
)
def test_real_workbook_second_upload_is_all_unchanged(app, client):
    app.config["MATERIAL_TAG_ISSUANCE_ENABLED"] = True
    user_id, station_id = _identity(app)
    _authenticate(client, user_id, station_id)
    workbook_bytes = REAL_WORKBOOK.read_bytes()
    _upload(client, workbook_bytes)
    client.post("/master-data/materials/import/1/apply")
    second = _upload(
        client,
        workbook_bytes,
        idempotency_key="22222222-2222-4222-8222-222222222222",
    )
    assert second.headers["Location"].endswith("/master-data/materials/import/2/preview")
    with app.app_context():
        batch = db.session.get(MaterialImportBatch, 2)
        assert (
            batch.inserted_count,
            batch.updated_count,
            batch.unchanged_count,
            batch.rejected_count,
        ) == (0, 0, 329, 0)
        assert Material.query.count() == 329
    client.post("/master-data/materials/import/2/apply")
    with app.app_context():
        assert Material.query.count() == 329
        assert AuditLog.query.filter_by(event_type="MATERIAL_IMPORT_APPLIED").count() == 2


def test_result_pages_are_paginated_and_route_methods_are_restricted(app, client):
    app.config["MATERIAL_TAG_ISSUANCE_ENABLED"] = True
    user_id, station_id = _identity(app)
    _authenticate(client, user_id, station_id)
    rows = [(f"MAT-{index:03d}", "MAT", f"Material {index}") for index in range(101)]
    response = _upload(client, _workbook_bytes(rows))
    page = client.get(response.headers["Location"])
    assert page.status_code == 200
    assert b"Next" in page.data
    assert b"MAT-099" in page.data
    assert b"MAT-100" not in page.data
    assert client.get("/master-data/materials/import/1/apply").status_code == 405
    client.post("/master-data/materials/import/1/apply")
    result = client.get("/master-data/materials/import/1/result?page=2")
    assert result.status_code == 200
    assert b"MAT-100" in result.data
