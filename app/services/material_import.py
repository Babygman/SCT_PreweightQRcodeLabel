import hashlib
import json
import unicodedata
import zipfile
from collections import Counter
from dataclasses import dataclass
from io import BytesIO

from defusedxml import ElementTree
from defusedxml.common import DefusedXmlException
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from sqlalchemy import func, select
from werkzeug.utils import secure_filename

from app.extensions import db
from app.models import AuditLog, Material, MaterialImportBatch, MaterialImportRow, utcnow

REQUIRED_HEADERS = ("ITEM CODE", "CATEGORY_NO", "NAME")
ALLOWED_CATEGORY = "MAT"
MAX_ZIP_ENTRIES = 1_000


class MaterialImportError(ValueError):
    """Raised when a Material Master workbook or import operation is invalid."""


@dataclass
class ParsedRow:
    row_number: int
    item_code: str | None
    category_no: str | None
    name: str | None
    result: str = "REJECTED"
    reason_code: str | None = None
    reason_detail: str | None = None


def _normalized(value, *, uppercase=False):
    if value is None:
        return None
    text_value = unicodedata.normalize("NFC", str(value)).strip()
    if not text_value:
        return None
    if any(unicodedata.category(character).startswith("C") for character in text_value):
        raise MaterialImportError("Workbook values must not contain control characters.")
    return text_value.upper() if uppercase else text_value


def _safe_filename(filename):
    basename = (filename or "").replace("\\", "/").rsplit("/", 1)[-1]
    cleaned = secure_filename(basename)
    return cleaned[:255] or "material-master.xlsx"


def _xml_local_name(element):
    return element.tag.rsplit("}", 1)[-1]


def _validate_xlsx_container(file_bytes, maximum_uncompressed_bytes):
    if not file_bytes.startswith(b"PK"):
        raise MaterialImportError("The uploaded file is not a valid .xlsx workbook.")
    try:
        with zipfile.ZipFile(BytesIO(file_bytes)) as archive:
            members = archive.infolist()
            if len(members) > MAX_ZIP_ENTRIES:
                raise MaterialImportError("The workbook contains too many internal files.")
            if any(
                member.filename.startswith(("/", "\\"))
                or ".." in member.filename.replace("\\", "/").split("/")
                for member in members
            ):
                raise MaterialImportError("The workbook contains an unsafe internal path.")
            if sum(member.file_size for member in members) > maximum_uncompressed_bytes:
                raise MaterialImportError("The workbook expands beyond the permitted size.")
            names = {member.filename for member in members}
            if "[Content_Types].xml" not in names or "xl/workbook.xml" not in names:
                raise MaterialImportError("The uploaded file is not a valid .xlsx workbook.")
            if any(name.lower().endswith("vbaproject.bin") for name in names):
                raise MaterialImportError("Macro-enabled workbooks are not accepted.")

            xml_documents = {}
            for member in members:
                if member.filename.lower().endswith((".xml", ".rels")):
                    try:
                        xml_documents[member.filename] = ElementTree.fromstring(
                            archive.read(member)
                        )
                    except (DefusedXmlException, ElementTree.ParseError, ValueError) as exc:
                        raise MaterialImportError(
                            "The workbook contains malformed or unsafe XML."
                        ) from exc

            if any(
                element.attrib.get("TargetMode", "").lower() == "external"
                for root in xml_documents.values()
                for element in root.iter()
                if _xml_local_name(element) == "Relationship"
            ):
                raise MaterialImportError("Workbooks containing external links are not accepted.")

            worksheet_xml = xml_documents.get("xl/worksheets/sheet1.xml")
            if worksheet_xml is not None and any(
                _xml_local_name(element) == "mergeCell" for element in worksheet_xml.iter()
            ):
                raise MaterialImportError("Merged cells are not accepted in Sheet1.")
    except (zipfile.BadZipFile, OSError, RuntimeError) as exc:
        raise MaterialImportError("The uploaded file is not a valid .xlsx workbook.") from exc


def parse_material_workbook(
    file_bytes,
    *,
    maximum_bytes,
    maximum_rows,
    maximum_uncompressed_bytes,
):
    if not file_bytes:
        raise MaterialImportError("The uploaded workbook is empty.")
    if len(file_bytes) > maximum_bytes:
        raise MaterialImportError("The workbook exceeds the permitted upload size.")
    _validate_xlsx_container(file_bytes, maximum_uncompressed_bytes)

    try:
        workbook = load_workbook(
            BytesIO(file_bytes), read_only=True, data_only=False, keep_links=False
        )
    except (
        InvalidFileException,
        KeyError,
        OSError,
        TypeError,
        ValueError,
        zipfile.BadZipFile,
    ) as exc:
        raise MaterialImportError(
            "The uploaded file could not be read as an .xlsx workbook."
        ) from exc

    try:
        if workbook.sheetnames != ["Sheet1"]:
            raise MaterialImportError("The workbook must contain only the Sheet1 worksheet.")
        worksheet = workbook["Sheet1"]
        worksheet_rows = list(worksheet.iter_rows())
        if not worksheet_rows:
            raise MaterialImportError("The Sheet1 worksheet is empty.")

        header_cells = list(worksheet_rows[0])
        while header_cells and _normalized(header_cells[-1].value) is None:
            header_cells.pop()
        if any(cell.data_type == "f" for cell in header_cells):
            raise MaterialImportError("Header cells must not contain formulas.")
        headers = tuple(_normalized(cell.value, uppercase=True) for cell in header_cells)
        if headers != REQUIRED_HEADERS:
            raise MaterialImportError(
                "Sheet1 headers must be exactly: ITEM CODE, CATEGORY_NO, NAME."
            )

        data_rows = [list(row[:3]) for row in worksheet_rows[1:]]
        while data_rows and all(
            cell.value is None or not str(cell.value).strip() for cell in data_rows[-1]
        ):
            data_rows.pop()
        if not data_rows:
            raise MaterialImportError("The workbook contains no Material rows.")
        if len(data_rows) > maximum_rows:
            raise MaterialImportError("The workbook contains more rows than permitted.")

        parsed = []
        for row_number, cells in enumerate(data_rows, start=2):
            values = [cell.value for cell in cells]
            normalized_values = []
            control_error = None
            for value, uppercase in zip(values, (True, True, False), strict=True):
                try:
                    normalized_values.append(_normalized(value, uppercase=uppercase))
                except MaterialImportError as exc:
                    normalized_values.append(None)
                    control_error = str(exc)
            code, category, name = normalized_values
            row = ParsedRow(row_number, code, category, name)
            if control_error:
                row.reason_code = "CONTROL_CHARACTER_NOT_ALLOWED"
                row.reason_detail = control_error
            elif any(cell.data_type == "f" for cell in cells):
                row.reason_code = "FORMULA_NOT_ALLOWED"
                row.reason_detail = "Required cells must contain values, not formulas."
            elif code is None or category is None or name is None:
                row.reason_code = "REQUIRED_VALUE_MISSING"
                row.reason_detail = "ITEM CODE, CATEGORY_NO, and NAME are required."
            elif len(code) > 50 or len(category) > 30 or len(name) > 200:
                row.reason_code = "VALUE_TOO_LONG"
                row.reason_detail = "One or more values exceed the permitted field length."
            elif category != ALLOWED_CATEGORY:
                row.reason_code = "CATEGORY_NOT_ALLOWED"
                row.reason_detail = "CATEGORY_NO must be MAT."
            parsed.append(row)

        duplicates = {
            code
            for code, count in Counter(row.item_code for row in parsed if row.item_code).items()
            if count > 1
        }
        for row in parsed:
            if row.item_code in duplicates:
                row.result = "REJECTED"
                row.reason_code = "DUPLICATE_ITEM_CODE"
                row.reason_detail = "ITEM CODE occurs more than once in this workbook."
        return parsed
    finally:
        workbook.close()


def _add_audit(event_type, batch, user_id, station_id):
    detail = json.dumps(
        {
            "batch_id": batch.id,
            "event": event_type,
            "filename": batch.original_filename,
            "file_sha256": batch.file_sha256,
            "status": batch.status,
            "counts": {
                "total": batch.total_rows,
                "insert": batch.inserted_count,
                "update": batch.updated_count,
                "unchanged": batch.unchanged_count,
                "rejected": batch.rejected_count,
            },
        },
        ensure_ascii=True,
        separators=(",", ":"),
    )
    audit = AuditLog(
        event_type=event_type,
        entity_type="MATERIAL_IMPORT_BATCH",
        entity_id=str(batch.id),
        user_id=user_id,
        station_id=station_id,
        occurred_at_utc=utcnow(),
        detail=detail,
    )
    if db.session.get_bind().dialect.name == "sqlite":
        audit.id = db.session.scalar(select(func.coalesce(func.max(AuditLog.id), 0) + 1))
    db.session.add(audit)


def _set_counts(batch):
    counts = Counter(row.result for row in batch.rows)
    batch.total_rows = len(batch.rows)
    batch.inserted_count = counts["INSERT"]
    batch.updated_count = counts["UPDATE"]
    batch.unchanged_count = counts["UNCHANGED"]
    batch.rejected_count = counts["REJECTED"]


def create_material_import_preview(
    *,
    file_bytes,
    filename,
    idempotency_key,
    user_id,
    station_id,
    maximum_bytes,
    maximum_rows,
    maximum_uncompressed_bytes,
):
    existing_batch = db.session.scalar(
        select(MaterialImportBatch).where(MaterialImportBatch.idempotency_key == idempotency_key)
    )
    if existing_batch is not None:
        return existing_batch

    file_hash = hashlib.sha256(file_bytes).hexdigest()
    batch = MaterialImportBatch(
        original_filename=_safe_filename(filename),
        file_sha256=file_hash,
        status="PREVIEWED",
        total_rows=0,
        uploaded_by_user_id=user_id,
        uploaded_at_utc=utcnow(),
        idempotency_key=idempotency_key,
    )
    db.session.add(batch)
    try:
        parsed_rows = parse_material_workbook(
            file_bytes,
            maximum_bytes=maximum_bytes,
            maximum_rows=maximum_rows,
            maximum_uncompressed_bytes=maximum_uncompressed_bytes,
        )
    except MaterialImportError as exc:
        batch.status = "FAILED"
        batch.error_summary = str(exc)
        db.session.flush()
        _add_audit("MATERIAL_IMPORT_FAILED", batch, user_id, station_id)
        db.session.commit()
        return batch

    valid_codes = [
        row.item_code
        for row in parsed_rows
        if row.reason_code is None and row.item_code is not None
    ]
    existing_materials = {
        material.code: material
        for material in db.session.scalars(select(Material).where(Material.code.in_(valid_codes)))
    }
    for parsed in parsed_rows:
        if parsed.reason_code is None:
            material = existing_materials.get(parsed.item_code)
            if material is None:
                parsed.result = "INSERT"
            elif material.name == parsed.name and material.source_category_no == parsed.category_no:
                parsed.result = "UNCHANGED"
            else:
                parsed.result = "UPDATE"
        batch.rows.append(
            MaterialImportRow(
                row_number=parsed.row_number,
                item_code_normalized=parsed.item_code,
                category_no_normalized=parsed.category_no,
                name_normalized=parsed.name,
                result=parsed.result,
                reason_code=parsed.reason_code,
                reason_detail=parsed.reason_detail,
            )
        )
    _set_counts(batch)
    db.session.flush()
    _add_audit("MATERIAL_IMPORT_PREVIEWED", batch, user_id, station_id)
    db.session.commit()
    return batch


def _batch_for_apply_statement(batch_id):
    return (
        select(MaterialImportBatch)
        .where(MaterialImportBatch.id == batch_id)
        .with_hint(MaterialImportBatch, "WITH (UPDLOCK, HOLDLOCK)", dialect_name="mssql")
    )


def _validate_persisted_preview(batch, rows):
    counts = Counter(row.result for row in rows)
    expected_counts = {
        "INSERT": batch.inserted_count,
        "UPDATE": batch.updated_count,
        "UNCHANGED": batch.unchanged_count,
        "REJECTED": batch.rejected_count,
    }
    if len(rows) != batch.total_rows or any(
        counts[result] != expected for result, expected in expected_counts.items()
    ):
        raise MaterialImportError("The persisted preview counts failed integrity validation.")
    if batch.rejected_count or counts["REJECTED"]:
        raise MaterialImportError(
            "This workbook contains rejected rows. Correct it and upload a new workbook."
        )
    if len({row.row_number for row in rows}) != len(rows) or any(
        row.row_number < 2 for row in rows
    ):
        raise MaterialImportError("The persisted Excel row identities are invalid.")

    codes = []
    for row in rows:
        if row.result not in {"INSERT", "UPDATE", "UNCHANGED"}:
            raise MaterialImportError("The persisted row classification is invalid.")
        try:
            code = _normalized(row.item_code_normalized, uppercase=True)
            category = _normalized(row.category_no_normalized, uppercase=True)
            name = _normalized(row.name_normalized)
        except MaterialImportError as exc:
            raise MaterialImportError("The persisted preview contains an invalid value.") from exc
        if (
            code is None
            or category != ALLOWED_CATEGORY
            or name is None
            or code != row.item_code_normalized
            or category != row.category_no_normalized
            or name != row.name_normalized
            or len(code) > 50
            or len(category) > 30
            or len(name) > 200
            or row.reason_code is not None
            or row.reason_detail is not None
        ):
            raise MaterialImportError("The persisted preview contains an invalid Material row.")
        codes.append(code)
    if len(codes) != len(set(codes)):
        raise MaterialImportError("The persisted preview contains duplicate Material codes.")


def apply_material_import(*, batch_id, user_id, station_id):
    batch = db.session.scalar(_batch_for_apply_statement(batch_id))
    if batch is None:
        raise MaterialImportError("Material import preview was not found.")
    if batch.status == "APPLIED":
        return batch
    if batch.status != "PREVIEWED":
        raise MaterialImportError("Only a valid preview can be applied.")

    rows = sorted(batch.rows, key=lambda row: row.row_number)
    _validate_persisted_preview(batch, rows)
    codes = [row.item_code_normalized for row in rows]
    existing_materials = {
        material.code: material
        for material in db.session.scalars(select(Material).where(Material.code.in_(codes)))
    }
    changed_at = utcnow()
    for row in rows:
        material = existing_materials.get(row.item_code_normalized)
        if material is None:
            material = Material(
                code=row.item_code_normalized,
                name=row.name_normalized,
                unit="kg",
                classification="GENERAL",
                source_category_no=row.category_no_normalized,
                updated_at_utc=changed_at,
                updated_by_user_id=user_id,
            )
            db.session.add(material)
            existing_materials[material.code] = material
            row.result = "INSERT"
        elif (
            material.name != row.name_normalized
            or material.source_category_no != row.category_no_normalized
        ):
            material.name = row.name_normalized
            material.source_category_no = row.category_no_normalized
            material.updated_at_utc = changed_at
            material.updated_by_user_id = user_id
            row.result = "UPDATE"
        else:
            row.result = "UNCHANGED"

    _set_counts(batch)
    batch.status = "APPLIED"
    batch.applied_by_user_id = user_id
    batch.applied_at_utc = changed_at
    _add_audit("MATERIAL_IMPORT_APPLIED", batch, user_id, station_id)
    db.session.commit()
    return batch
